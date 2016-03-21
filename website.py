from openpyxl import*
from flask import*

def shutdown_server():
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug Server')
    func()



def serv1(mnj):
        app = Flask(__name__)

        @app.route('/')
        def index():
            return render_template('index.html')

        @app.route('/login', methods = ['POST'])
        def login():
            if int(request.form['pass'])==1234 :
                PATH='attendence.xlsx'
                if os.path.isfile(PATH):
                    wb = load_workbook('attendence.xlsx')
                else:
                    wb=Workbook()
                ws = wb.active
                today_column=ws.max_column
                column1=ws.columns[0]
                for j in range(1,len(column1)):
                    exist=False
                    if column1[j].value == request.form['roll']:
                        value=j
                        exist=True
                        break
                    

                if exist:
                    ws.cell(row=ws.cell(column1[value].coordinate).row,column=today_column).value=1
                    wb.save("attendence.xlsx")
                    return render_template('success.html')
                else:
                    return render_template('error.html')
            else:
                 return render_template('wrongpass.html')

        @app.route('/shutdown')
        def shutdown():
            shutdown_server()
            return 'Server shutting down...'

        
        app.run(host='0.0.0.0',port='80')


            
