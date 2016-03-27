from Tkinter import*
import urllib2
from openpyxl import*
import os
from datetime import date
import shutil
import tkFileDialog
from website import*
from get_headcount import*

from ScrolledText import ScrolledText
from ttk import*
import thread
from multiprocessing import Process, Manager, Queue
from Queue import Empty
from decimal import Decimal, getcontext


DELAY1 = 80
DELAY2 = 20

# Queue must be global
q = Queue()


class Example(Frame):
  
    def __init__(self, parent):
        Frame.__init__(self, parent, name="frame")   
                 
        self.parent = parent
        self.initUI()

    def initUI(self):
        self.grid()
        
        

        self.labeltext1=StringVar()
        label1=Label(self,textvar=self.labeltext1)
        label1.grid(row=0,column=0)

        self.labeltext=StringVar()
        label=Label(self,textvar=self.labeltext)
        label.grid(row=0,column=2)

        button1=Button(self,text="START TODAY'S ATTEND.")
        button1.grid(row=3,column=0)

        button2=Button(self,text="EDITING LAST ATTEND.")
        button2.grid(row=3,column=2)

        button3=Button(self,text="START TAKING")
        button3.grid(row=4,column=0)

        button4=Button(self,text="STOP TAKING",command=self.func4)
        button4.grid(row=4,column=2)

        button5=Button(self,text="SEE TODAY'S ATTEND.")
        button5.grid(row=5,column=0)

        self.button6=Button(self,text="DOWNLOAD ATTEND.",command=self.func6)
        self.button6.grid(row=5,column=2)

        self.button7=Button(self,text="GET HEAD COUNT",command=self.func7)
        self.button7.grid(row=6,column=2)

        self.pbar = Progressbar(self, mode='indeterminate')        
        self.pbar.grid(row=7, column=1, sticky=W+E)     
        
        #self.txt = ScrolledText(self)  
        #self.txt.grid(row=8, column=0, rowspan=4, padx=10, pady=5,
         #   columnspan=5, sticky=E+W+S+N)
            

        button1.bind('<Button-1>',self.func1)
        button2.bind('<Button-1>',self.func2)
        button3.bind('<Button-1>',self.func3)
        #button4.bind('<Button-1>',self.func4)
        button5.bind('<Button-1>',self.func5)
        #self.button6.bind('<Button-1>',self.func6)
        self.button7.bind('<Button-1>',self.func7)

    def func1(self,event):
        PATH='attendence.xlsx'
        if os.path.isfile(PATH):
            wb = load_workbook('attendence.xlsx')
        else:
            wb=Workbook()

        ws = wb.active
        today_column=ws.max_column+1
        date1=date.today()
        ws.cell(row=1,column=today_column).value=date1
        for i in range(2,ws.max_row+1):
            ws.cell(row=i,column=today_column).value=0
        wb.save("attendence.xlsx")
        

    def func2(self,event):
        print('ALLOWED')


    def func3(self,event):
           thread.start_new_thread(serv1, ("Thread-1",) )

    

    def func4(self):
        response = urllib2.urlopen('http://localhost/shutdown')
        #thread.start_new_thread(self.afun, ("Thread-2",) )

    def func5(self,event):
            self.labeltext1.set('Loading...')
            PATH='attendence.xlsx'
            if os.path.isfile(PATH):
                wb = load_workbook('attendence.xlsx')
                ws = wb.active
                today_column=ws.max_column
                column1=ws.columns[today_column-1]
                sum=0
                for j in range(1,len(column1)):
                    sum=sum+column1[j].value
                self.labeltext1.set(sum)

    def func6(self):
        #self.withdraw() #use to hide tkinter window
        currdir = os.getcwd()
        tempdir = tkFileDialog.askdirectory(parent=self, initialdir=currdir, title='Please select a directory')
        if len(tempdir) > 0:
            shutil.copy('attendence.xlsx', tempdir)
        else:
            pass

    

    def func7(self,event):
        
        self.button7.config(state=DISABLED)
        #self.txt.delete("1.0", END)
        
        self.p1 = Process(target=head_cnt, args=(q,))
        self.p1.start()
        self.pbar.start(DELAY2)
        self.after(DELAY1, self.onGetValue)
        
       
    def onGetValue(self):
        
        if (self.p1.is_alive()):
            
            self.after(DELAY1, self.onGetValue)
            return
        else:    
        
           try:
                self.labeltext.set(q.get(0)) 
                #self.txt.insert('end', q.get(0))
                #self.txt.insert('end', "\n")
                self.pbar.stop()
                self.button7.config(state=NORMAL)

           except Empty:
                print("queue is empty")

def main():  
  
    root = Tk()
    root.geometry("400x200+300+300")
    app = Example(root)
    root.mainloop()  


if __name__ == '__main__':
    main()
