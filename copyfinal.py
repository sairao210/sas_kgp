from Tkinter import*
import thread
from openpyxl import*
import os
from datetime import date
from flask import*
from multiprocessing import Process
import shutil
import tkFileDialog


class appgui(Tk):
    def __init__(self):
        Tk.__init__(self)
        self.initialize()

    def initialize(self):
        self.grid()
        
        self.labeltext=StringVar()
        label=Label(self,textvar=self.labeltext,bg='white')
        label.grid(row=0,sticky='E')

        button1=Button(self,text="START TODAY'S ATTEND.")
        button1.grid(row=3,column=0)

        button2=Button(self,text="EDITING LAST ATTEND.")
        button2.grid(row=3,column=1)

        button3=Button(self,text="START TAKING")
        button3.grid(row=4,column=0)

        button4=Button(self,text="STOP TAKING")
        button4.grid(row=4,column=1)

        button5=Button(self,text="SEE TODAY'S ATTEND.")
        button5.grid(row=5,column=0)

        button6=Button(self,text="DOWNLOAD ATTEND.")
        button6.grid(row=5,column=1)


        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)

        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.grid_rowconfigure(3,weight=1)
        self.grid_rowconfigure(4,weight=1)
        self.grid_rowconfigure(5,weight=1)

    

        button1.bind('<Button-1>',self.func1)
        button2.bind('<Button-1>',self.func2)
        button3.bind('<Button-1>',self.func3)
        button4.bind('<Button-1>',self.func4)
        button5.bind('<Button-1>',self.func5)
        button6.bind('<Button-1>',self.func6)

    def func1(self,event):
        PATH='attendence.xlsx'
        if os.path.isfile(PATH):
            wb = load_workbook('attendence.xlsx')
        else:
            wb=Workbook()

        ws = wb.active
        today_column=ws.max_column+1
        date1=date.today()
        ws.cell(row=1,column=today_column).value=date1
        for i in range(2,ws.max_row+1):
            ws.cell(row=i,column=today_column).value=0
        wb.save("attendence.xlsx")
        

    def func2(self,event):
        print('ALLOWED')






    def serv1(self):
        app = Flask(__name__)

        @app.route('/')
        def index():
            return render_template('index.html')
 
        @app.route('/login', methods = ['POST'])
        def login():
            if int(request.form['pass'])==1234 :
                PATH='attendence.xlsx'
                if os.path.isfile(PATH):
                    wb = load_workbook('attendence.xlsx')
                else:
                    wb=Workbook()
                ws = wb.active
                today_column=ws.max_column
                column1=ws.columns[0]
                for j in range(1,len(column1)):
                    exist=False
                    if column1[j].value == request.form['roll']:
                        value=j
                        exist=True
                        break
                    

                if exist:
                    ws.cell(row=ws.cell(column1[value].coordinate).row,column=today_column).value=1
                    wb.save("attendence.xlsx")
                    return render_template('success.html')
                else:
                    return render_template('error.html')
            else:
                 return render_template('wrongpass.html')

        if __name__ == "__main__":
            app.run(host='0.0.0.0',port='80')









    def func3(self,event):
        global child
        child = Process(target = self.serv1 )
        



    def func4(self,event):
        child.terminate()

    def func5(self,event):
        #self.withdraw() #use to hide tkinter window
        currdir = os.getcwd()
        tempdir = tkFileDialog.askdirectory(parent=self, initialdir=currdir, title='Please select a directory')
        if len(tempdir) > 0:
            shutil.copy('attendence.xlsx', tempdir)
        else:
            pass
        print('hey')

    def func6(self,event):
        #self.withdraw() #use to hide tkinter window
        currdir = os.getcwd()
        tempdir = tkFileDialog.askdirectory(parent=self, initialdir=currdir, title='Please select a directory')
        if len(tempdir) > 0:
            shutil.copy('attendence.xlsx', tempdir)
        else:
            pass
        print('hey')
            

    
       
        

new= appgui()

new.mainloop()
